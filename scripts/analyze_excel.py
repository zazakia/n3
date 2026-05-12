
import openpyxl
wb = openpyxl.load_workbook('brayan Import migration cleanup.xlsx', data_only=True)
ws = wb.active

batch_rows = []
for r in range(1, ws.max_row + 1):
    v1 = ws.cell(r, 1).value
    if v1 and isinstance(v1, str) and 'batch' in v1.lower():
        batch_rows.append(r)

rows_out = []
for i, batch_row in enumerate(batch_rows):
    next_row = batch_rows[i+1] if i+1 < len(batch_rows) else ws.max_row+1
    data_start = batch_row + 3
    data_end = next_row - 1
    data_count = 0
    collector = ''
    days_val = ''
    for r in range(data_start, data_end+1):
        v1 = ws.cell(r, 1).value
        # If we hit an empty row or 'Total', stop counting data for this batch
        if v1 and isinstance(v1, str) and v1.lower().strip() == 'total':
            break
        if v1 and str(v1).strip():
            # Looks like a borrower row
            # Double check there is a borrower name
            if len(str(v1).strip()) > 3:
                data_count += 1
                if not collector:
                    collector = str(ws.cell(r, 6).value or '').strip()
                    days_val = str(ws.cell(r, 7).value or '').strip()
    
    freq_raw = str(ws.cell(batch_row, 10).value or '')
    batch_name = str(ws.cell(batch_row, 1).value).strip()
    if 'weekly' in batch_name.lower():
        freq = 'weekly'
    elif 'weekly' in freq_raw.lower():
        freq = 'weekly'
    else:
        freq = 'daily'
    is_reloan = 'reloan' in batch_name.lower()
    rows_out.append((batch_name, batch_row, data_count, collector, days_val, freq, is_reloan))

print("Batch,Row,Borrowers,Collector,Days,Freq,IsReloan")
total_borrowers = 0
for r in rows_out:
    print(f"{r[0]},{r[1]},{r[2]},{r[3]},{r[4]},{r[5]},{r[6]}")
    total_borrowers += r[2]

print(f"TOTAL borrower records: {total_borrowers}")
unique_collectors = list(set(r[3] for r in rows_out if r[3]))
print(f"Collectors: {unique_collectors}")
