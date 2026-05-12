
import openpyxl
from collections import Counter

wb = openpyxl.load_workbook('brayan Import migration cleanup.xlsx', data_only=True)
ws = wb.active

batch_rows = []
for r in range(1, ws.max_row + 1):
    v1 = ws.cell(r, 1).value
    if v1 and isinstance(v1, str) and 'batch' in v1.lower():
        batch_rows.append(r)

borrower_colors = Counter()
active_count = 0
paid_count = 0
paid_details = []

for i, batch_row in enumerate(batch_rows):
    next_row = batch_rows[i+1] if i+1 < len(batch_rows) else ws.max_row+1
    data_start = batch_row + 3
    data_end = next_row - 1
    
    for r in range(data_start, data_end+1):
        cell = ws.cell(r, 1)
        v1 = cell.value
        if v1 and isinstance(v1, str) and v1.lower().strip() == 'total':
            break
        if v1 and str(v1).strip():
            if len(str(v1).strip()) > 3:
                # This is a borrower row. Check color.
                color = 'None'
                is_orange = False
                
                if cell.fill and hasattr(cell.fill, 'fgColor') and hasattr(cell.fill.fgColor, 'rgb'):
                    rgb = cell.fill.fgColor.rgb
                    if type(rgb) == str:
                        color = rgb
                        # FFC000 is standard orange, ED7D31 is another orange, etc.
                        # Anything that's not 00000000 or None and has a solid pattern we'll count as "filled"
                        if color not in ['00000000', 'None', 'FFFFFFFF'] and cell.fill.patternType == 'solid':
                            is_orange = True
                
                if is_orange:
                    paid_count += 1
                    paid_details.append((r, str(v1).strip(), color))
                else:
                    active_count += 1
                
                borrower_colors[color] += 1

print(f"Total Active (White/Unfilled): {active_count}")
print(f"Total Paid (Orange): {paid_count}")
print(f"Total Scanned: {active_count + paid_count}")
print("\nColor Distribution (Hex Codes Found):")
for c, count in borrower_colors.most_common():
    print(f"  {c}: {count}")

print("\nSample of Paid Borrowers:")
for p in paid_details[:10]:
    print(f"  Row {p[0]}: {p[1]} (Color: {p[2]})")

