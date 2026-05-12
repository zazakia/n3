
import openpyxl

wb = openpyxl.load_workbook('brayan Import migration cleanup.xlsx', data_only=True)
ws = wb.active

paid_loans = 0
active_loans = 0
borrower_names = set()
collectors = set()

for r in range(4, ws.max_row+1):
    cell = ws.cell(r, 1)
    val = cell.value
    
    if val and isinstance(val, str) and not val.lower().startswith('batch') and not val.lower().startswith('total') and not val.lower().startswith('grand'):
        if len(val.strip()) > 3:
            # We have a valid borrower row
            # Get collector name (Column F / 6)
            collector_val = ws.cell(r, 6).value
            if collector_val and isinstance(collector_val, str) and len(collector_val.strip()) > 1:
                collectors.add(collector_val.strip())
            
            borrower_names.add(val.strip())
            
            # Check cell color for paid vs active
            is_paid = False
            if cell.fill and hasattr(cell.fill, 'fgColor'):
                fg = cell.fill.fgColor
                theme = getattr(fg, 'theme', None)
                tint = getattr(fg, 'tint', None)
                if theme == 5 and tint is not None and abs(float(tint) - 0.6) < 0.1:
                    is_paid = True
            
            if is_paid:
                paid_loans += 1
            else:
                active_loans += 1

print(f"Total Loans Detected: {paid_loans + active_loans}")
print(f"  - Active Loans (White): {active_loans}")
print(f"  - Paid Loans (Orange): {paid_loans}")
print(f"  - Total Distinct Borrowers (Unique Names): {len(borrower_names)}")
print(f"  - Total Unique Collectors: {len(collectors)}")

print("\nCollectors List:")
for c in sorted(list(collectors)):
    print(f"  - {c}")

