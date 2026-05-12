
import openpyxl

wb = openpyxl.load_workbook('brayan Import migration cleanup.xlsx', data_only=True)
ws = wb.active

paid_loans = 0
active_loans = 0
borrower_names = set()
collectors = set()

# Safe list of keywords to ignore in Column A
ignore_kw = ['batch', 'total', 'grand total', 'name of client', 'monthly', 'weekly']

for r in range(4, ws.max_row+1):
    cell = ws.cell(r, 1)
    val = cell.value
    
    if val and isinstance(val, str):
        v_lower = val.lower().strip()
        
        # Skip header rows
        should_skip = False
        for kw in ignore_kw:
            if v_lower.startswith(kw):
                should_skip = True
                break
        
        if not should_skip and len(v_lower) > 3:
            # Check cell color for paid vs active
            is_paid = False
            if cell.fill and hasattr(cell.fill, 'fgColor'):
                fg = cell.fill.fgColor
                theme = getattr(fg, 'theme', None)
                tint = getattr(fg, 'tint', None)
                if theme == 5 and tint is not None and abs(float(tint) - 0.6) < 0.1:
                    is_paid = True
            
            # Record collector
            collector_val = ws.cell(r, 6).value
            if collector_val and isinstance(collector_val, str) and len(collector_val.strip()) > 3:
                c_str = collector_val.strip().lower()
                if not 'collector' in c_str:  # Ignore headers saying 'Collector 1'
                    collectors.add(collector_val.strip())

            borrower_names.add(val.strip())
            
            if is_paid:
                paid_loans += 1
            else:
                active_loans += 1

print("--- DEFINITIVE COUNTS ---")
print(f"Total Loans: {paid_loans + active_loans}")
print(f"  - Active Loans (White): {active_loans}")
print(f"  - Paid Loans (Orange): {paid_loans}")
print(f"  - Unique Borrower Names: {len(borrower_names)}")
print(f"  - Unique Collector Names: {len(collectors)}")

print("\nReal Collectors Found:")
for c in sorted(list(collectors)):
    print(f"  - {c}")
