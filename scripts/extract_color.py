
import openpyxl

wb = openpyxl.load_workbook('brayan Import migration cleanup.xlsx', data_only=True)
ws = wb.active

for r in range(1, ws.max_row+1):
    cell = ws.cell(r, 1)
    if cell.value and isinstance(cell.value, str) and 'anita d. pepito' in cell.value.lower():
        print(f'Found {cell.value} at Row {r}')
        if cell.fill and hasattr(cell.fill, 'fgColor'):
            fg = cell.fill.fgColor
            print(f"  rgb: {getattr(fg, 'rgb', 'N/A')}")
            print(f"  theme: {getattr(fg, 'theme', 'N/A')}")
            print(f"  tint: {getattr(fg, 'tint', 'N/A')}")
            print(f"  type: {getattr(fg, 'type', 'N/A')}")
        else:
            print("  No fill found on Name cell.")
